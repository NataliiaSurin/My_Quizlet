import openpyxl as op
import random
import telebot
from telebot import types
from urllib.request import urlretrieve
import os
# os.remove("Courses.xlsx")

TOKEN = "7379053950:AAEDhfRYutZmdcFxXMOut8ZjbSYmngudYSA"  # Token for Telegram Bot
bot = telebot.TeleBot(TOKEN)
user_counters = {}  # Initiation dictionary for score counting inside bot.callback_query_handler
global total_rows


# Telebot /start command
@bot.message_handler(commands=['start'])
def start(message):
    user_id = message.from_user.id
    user_counters[user_id] = 0
    markup = types.InlineKeyboardMarkup()
    markup.add(types.InlineKeyboardButton("Go to the Cambridge dictionary website", url='https://dictionary.cambridge.org/plus/myWordlists'))
    bot.send_message(message.chat.id, f'Hello, {message.from_user.first_name} ! Upload Exel file with your words list from Cambridge dictionary', reply_markup=markup)


# Exel file uploading handler
@bot.message_handler(content_types=['document'])
def get_file(message):
    if message.document.mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":  # Checking file type
        markup2 = types.InlineKeyboardMarkup()
        markup2.add(types.InlineKeyboardButton("Start game", callback_data='start_game'))
        bot.reply_to(message, "Your words list uploaded successfully", reply_markup=markup2)
        bot.get_file(message.document.file_id)
        url = str(bot.get_file_url(message.document.file_id))
        file_name = message.document.file_name
        urlretrieve(url, file_name)
        df = op.load_workbook(file_name)
        data = df.active
        global total_rows
        total_rows = data.max_row
        global words, definitions, index_list
        words = []
        definitions = []

        # Creating list with words
        for row_number in range(3, total_rows + 1):
            cell_obj = data.cell(row=row_number, column=1)
            words.append(cell_obj.value)

        index_list = list(range(len(words)))    # Creating list of indices

        # Creating list with definitions
        for row_number in range(3, total_rows + 1):
            cell_obj = data.cell(row=row_number, column=3)
            definitions.append(cell_obj.value)

    else:
        bot.send_message(message.chat.id, "Not correct file format. Supposed *.xslx. Please try again.")


# Choosing 4 random rows for game round
def randomize(index_list):
    global random_row_1, random_row_2, random_row_3, random_row_4, random_row_main
    index_list_temporary = index_list.copy()  # Creating temporary list of indices
    random_row_1 = random.choice(index_list_temporary)
    index_list_temporary.remove(random_row_1)  # Removing of already chose index from temporary list of indices, so then they will not duplicate
    random_row_2 = random.choice(index_list_temporary)
    index_list_temporary.remove(random_row_2)
    random_row_3 = random.choice(index_list_temporary)
    index_list_temporary.remove(random_row_3)
    random_row_4 = random.choice(index_list_temporary)
    random_row_main = random.choice([random_row_1, random_row_2, random_row_3, random_row_4])


# Function for creating new game round with 4 answer variants buttons
def send_question(callback):
    randomize(index_list)
    markup3 = types.InlineKeyboardMarkup()
    markup3.add(types.InlineKeyboardButton(f"{words[random_row_1]}", callback_data=str(random_row_1)))
    markup3.add(types.InlineKeyboardButton(f"{words[random_row_2]}", callback_data=str(random_row_2)))
    markup3.add(types.InlineKeyboardButton(f"{words[random_row_3]}", callback_data=str(random_row_3)))
    markup3.add(types.InlineKeyboardButton(f"{words[random_row_4]}", callback_data=str(random_row_4)))
    kb = types.ReplyKeyboardMarkup(resize_keyboard=True)
    button_1 = types.KeyboardButton(text='Stop')
    kb.add(button_1)
    bot.send_message(callback.message.chat.id, f'Choose the right word that best matches to this definition:', reply_markup=kb)
    bot.send_message(callback.message.chat.id, f'<b>{definitions[random_row_main]}</b>', parse_mode='html', reply_markup=markup3)


# Analysis pressed buttons, checking for answer correctness, score counting
@bot.callback_query_handler(func=lambda callback: True)
def callback_message(callback):
    user_id = callback.from_user.id
    print(user_counters[user_id])
    if callback.data == 'start_game':
        send_question(callback)
        bot.answer_callback_query(callback.id)
        print(callback.data)
    elif callback.data == str(random_row_main):
        if user_id in user_counters:
            user_counters[user_id] += 1  # Score counting
        print(f'Correct, you score now {user_counters[user_id]}')
        bot.answer_callback_query(callback.id)
        bot.send_message(callback.message.chat.id, f'<b>Correct</b>, you score now <b>{user_counters[user_id]}</b>', parse_mode='html')
        send_question(callback)
    else:
        print(f'Wrong, you score now {user_counters[user_id]}')
        bot.answer_callback_query(callback.id)
        bot.send_message(callback.message.chat.id, f'<b>Wrong</b>, you score now <b>{user_counters[user_id]}</b>', parse_mode='html')
        send_question(callback)


bot.polling(non_stop=True)
